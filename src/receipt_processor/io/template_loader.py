"""Load model template schema and formatting hints."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class TemplateHints:
    """Presentation hints inferred from example files."""

    date_output_format: str = "%Y-%m-%d"
    currency_symbol: str = ""
    date_column: str = ""
    description_column: str = ""
    amount_column: str = ""
    receipt_amount_column: str = ""


def _read_csv_rows(path: Path, limit: int | None = None) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[dict[str, str]] = []
        for idx, row in enumerate(reader):
            rows.append({k: (v or "") for k, v in row.items()})
            if limit is not None and idx + 1 >= limit:
                break
        return rows


def _read_xlsx_rows(path: Path, limit: int | None = None) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    values = list(sheet.values)
    if not values:
        return []

    header = [str(value or "").strip() for value in values[0]]
    rows: list[dict[str, str]] = []
    for idx, row_values in enumerate(values[1:]):
        row_dict: dict[str, str] = {}
        for col_idx, col_name in enumerate(header):
            value = row_values[col_idx] if col_idx < len(row_values) else ""
            row_dict[col_name] = "" if value is None else str(value)
        rows.append(row_dict)
        if limit is not None and idx + 1 >= limit:
            break
    return rows


def _read_rows(path: Path, limit: int | None = None) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path, limit=limit)
    if suffix == ".xlsx":
        return _read_xlsx_rows(path, limit=limit)
    raise ValueError(f"Unsupported template format: {path}")


def load_model_columns(model_file: Path) -> list[str]:
    """Load column names from a model template."""
    rows = _read_rows(model_file, limit=1)
    if rows:
        return list(rows[0].keys())

    # Empty template still has header row in CSV/XLSX. Fallback read by file type.
    if model_file.suffix.lower() == ".csv":
        with model_file.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader, [])
        return [str(column) for column in header]

    if model_file.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("openpyxl is required to read .xlsx templates") from exc
        workbook = load_workbook(model_file, read_only=True, data_only=True)
        sheet = workbook.active
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(column or "") for column in header_values]

    raise ValueError(f"Unsupported model format: {model_file}")


def load_model_rows(model_file: Path) -> list[dict[str, str]]:
    """Load data rows (excluding header) from model template."""
    return _read_rows(model_file)


def _normalize_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()


def _find_column(columns: list[str], candidates: tuple[str, ...]) -> str:
    normalized_to_original = { _normalize_label(col): col for col in columns }
    for candidate in candidates:
        for norm, original in normalized_to_original.items():
            if candidate in norm:
                return original
    return ""


def _infer_date_format(sample_value: str) -> str:
    value = sample_value.strip()
    if re.fullmatch(r"\d{4}\s\d{2}\s\d{2}", value):
        return "%Y %m %d"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return "%Y-%m-%d"
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", value):
        return "%m/%d/%Y"
    return "%Y-%m-%d"


def _infer_currency_symbol(sample_value: str) -> str:
    value = sample_value.strip()
    if value.startswith("$"):
        return "$"
    return ""


def _first_non_empty(rows: list[dict[str, str]], column: str) -> str:
    for row in rows:
        value = (row.get(column) or "").strip()
        if value:
            return value
    return ""


def infer_template_hints(model_file: Path, example_file: Path) -> TemplateHints:
    """Infer formatting and column-role hints from model + example files."""
    model_columns = load_model_columns(model_file)
    example_rows = _read_rows(example_file, limit=50)

    date_col = _find_column(model_columns, ("date",))
    description_col = _find_column(model_columns, ("description", "memo", "details"))
    amount_col = _find_column(model_columns, ("amt claimed", "amount", "total"))
    receipt_amount_col = _find_column(model_columns, ("receipt amt", "receipt amount"))

    date_sample = _first_non_empty(example_rows, date_col) if date_col else ""
    amount_sample = _first_non_empty(example_rows, amount_col) if amount_col else ""

    return TemplateHints(
        date_output_format=_infer_date_format(date_sample) if date_sample else "%Y-%m-%d",
        currency_symbol=_infer_currency_symbol(amount_sample) if amount_sample else "",
        date_column=date_col,
        description_column=description_col,
        amount_column=amount_col,
        receipt_amount_column=receipt_amount_col,
    )
